from flask import Flask, render_template, request, flash ,session,send_file
from database import users, db_init, db
import re
import csv
from werkzeug.utils import secure_filename
from csrlc import replace_csv
import os
import pytesseract
from io import TextIOWrapper
import xlrd
import openpyxl 
from openpyxl.utils.cell import get_column_letter
import os
#import urllib.urlopen
from urllib.request import urlopen
from json import load
import json
import requests




#from xlrd import open_workbook

import base64 

app = Flask(__name__)
app.secret_key="1258suy"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///users.sqlite"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db_init(app)
@app.route('/')
def index():
   return render_template('index.html')

@app.route('/typein.html',methods=['GET','POST'])
def typein():
   #if request.method=='POST':
   
      #txt=request.form['text']
      
   #else:
   return render_template('typein.html')

 
   return render_template('typein.html')

@app.route('/xlsx.html')
def xlsx():
   return render_template('xlsx.html')

@app.route('/csv.html')
def csv():
   return render_template('csv.html')

@app.route('/text.html')
def text():
   return render_template('text.html')

@app.route('/typein-find.html',methods=['GET','POST'])
def find():
   if request.method=='POST':
      check=request.form['find']
      txt=request.form['text']
      
      #a = users( txt, check,"")
      x=re.search(check, txt)
      #print(x.span())
      #db.session.add(a)
      #db.session.commit()
      
      if x:
         flash(f"Word is found at the position {x.span()}")
         #print('Y')
         #return render_template('typein-find.html')
      else:
         flash("Word is not found")

         #return render_template('typein-find.html')

   return render_template("typein-find.html")

@app.route('/typein-replace.html',methods=['GET','POST'])
def replace():
   if request.method=='POST':
      check1=request.form['find']
      txt1=request.form['text']
      rlc=request.form['replace']
      new_string = re.sub(check1,rlc,txt1)
      flash(new_string)

   return render_template("typein-replace.html")

@app.route('/typein-number.html',methods=['GET','POST'])
def number():
   if request.method=='POST':
      txt1=request.form['text']
      pattern = '\d+'
      result = re.findall(pattern, txt1) 
      flash(result)
      

   return render_template("typein-number.html")

@app.route('/typein-https.html',methods=['GET','POST'])
def https():
   if request.method=='POST':
      txt1=request.form['text']
      obj1 = re.findall('(\w+)://',txt1) 
      flash(f'Protocol:{obj1}') 
      obj2 = re.findall('://([\w\-\.]+)',txt1) 
      flash(f'Host:{obj2}')
   return render_template("typein-https.html")

@app.route('/typein-ip.html',methods=['GET','POST'])
def ip():
   if request.method=='POST':
      txt1=request.form['text']
      regex = '''^(25[0-5]|2[0-4][0-9]|[0-1]?[0-9][0-9]?)\.( 
            25[0-5]|2[0-4][0-9]|[0-1]?[0-9][0-9]?)\.( 
            25[0-5]|2[0-4][0-9]|[0-1]?[0-9][0-9]?)\.( 
            25[0-5]|2[0-4][0-9]|[0-1]?[0-9][0-9]?)$'''
      if(re.search(regex, txt1)): 
         flash("Valid Ip address")  
      else:
         flash("Invalid Ip address")
   return render_template("typein-ip.html")

@app.route('/typein-mail.html',methods=['GET','POST'])
def mail():
   if request.method=='POST':
      txt1=request.form['text']
      regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
      if(re.search(regex,txt1)):
         flash("Valid Email")  
      else:
         flash("Invalid Email")
   return render_template("typein-mail.html")

@app.route('/typein-quotedword.html',methods=['GET','POST'])
def quotedword():
   if request.method=='POST':
      txt1=request.form['text']
      flash(re.findall(r'"(.*?)"', txt1))
   return render_template("typein-quotedword.html")
@app.route('/IP_location.html',methods=['GET','POST'])
def ip_location():
   if request.method=='POST':
      while True:
         txt1=request.form['text']
         url="http://ip-api.com/json/"
         response=urlopen(url+txt1)
         data=response.read()
         values=json.loads(data)
         flash(" IP: " + values['query'])
         flash(" City: " + values['city'])
         flash(" ISP: " + values['isp'])
         flash(" Country: " + values['country'])
         flash(" Region: " + values['region'])
         flash(" Time zone: " + values['timezone'])
         break
   return render_template("IP_location.html")

@app.route('/mobile_location.html',methods=['GET','POST'])
def mobile_location():
   import phonenumbers
   from phonenumbers import timezone
   from phonenumbers import geocoder
   from phonenumbers import carrier
   if request.method=='POST': 
      countrycode = request.form['text2']
      phonenumber = request.form['text']
      num = countrycode+phonenumber
      number = phonenumbers.parse(num)
      flash(geocoder.description_for_number(number,'en',region="GB"))
      flash(carrier.name_for_number(number, 'en'))
      timeZone = timezone.time_zones_for_number(number)
      flash(timeZone)

   return render_template("mobile_location.html")
#for text
@app.route('/text-find.html',methods=['GET','POST'])
def text_find():
   return render_template("text-find.html")


@app.route('/text-replace.html',methods=['GET','POST'])
def text_replace():
   return render_template("text-replace.html")

#for csv
@app.route('/csv-find.html',methods=['GET','POST'])

def cv_find():
   import csv
   
   if request.method=='POST':
      ls=[]
      find=request.form['find']
      text=request.files['actual-btn']
      #filename=text.split(".")
      text = TextIOWrapper(text, encoding='utf-8')
      

     


      
      cv_reader1= csv.reader(text)
      for line in cv_reader1:
         for word in line:
            a=word.split(';')
            ls=ls+a
      for item in ls:
         if item in find:
            flash('Found')
            break
      else:
         flash('Not Found')
      

   
   return render_template("csv-find.html")
@app.route('/csv-replace.html',methods=['GET','POST'])
def csv_replace():
   import csv
   text = b""
   if request.method=='POST':
      org_wrd=request.form['Replace']
      rlc_wrd=request.form['nwwrd']
      txt= request.files['actual-btn']
      txt = TextIOWrapper(txt, encoding='utf-8')
      csv_reader=csv.reader(txt)
     
      text = ''.join([str(i) for i in csv_reader])  
      
      text = text.replace('{}'.format(org_wrd),'{}'.format(rlc_wrd)) 
      x = open("output.csv","w")
      x.writelines(text)
      x.close()

      '''text = bytes(text, 'utf-8')'''
      text = text.encode('utf-8')
      
   return render_template("csv-replace.html",op=text.decode('utf-8'))

   #return render_template("csv-replace.html",op=base64.b64encode(text).decode('utf-8'))

@app.route('/xlsx-find.html',methods=['GET','POST'])
def xlx_find():
   import xlrd
   #from xlrd import open_workbook
   if request.method=='POST':
      find=request.form['find']
      txt=request.files['actual-btn']
      #filename=text.split(".")
      #txt = TextIOWrapper(txt, encoding='utf-8')
      #txt.save(os.path.join( secure_filename(txt.filename))
      filename = secure_filename(txt.filename)
      new_path = os.path.abspath(filename)
      w = xlrd.open_workbook(new_path)
      sheet = w.sheet_by_index(0)
      count=0
      
      for i in range(0,sheet.nrows):
         for j in range(0,sheet.ncols):
            if sheet.cell_value(i,j) == find:
               flash(f"Word Found at the position ,{i}*{j}")
               count=count+1
      if count==0:
         flash("Word is not Found")

   return render_template("xlsx-find.html")

from flask import  send_file

@app.route('/download/<filename>',methods=['GET'])
def download(filename):
   return send_file(filename, as_attachment=True)

@app.route('/xlsx-replace.html',methods=['GET','POST'])
def xlx_replace():
   filename = "" 
   if request.method=='POST':
      org_wrd=request.form['Replace']
      rlc_wrd=request.form['nwwrd']
      txt= request.files['actual-btn']
      #txt = TextIOWrapper(txt, encoding='utf-8')
      filename = secure_filename(txt.filename)
      new_path = os.path.abspath(filename)
      workbook = openpyxl.load_workbook(new_path) 
      
      worksheet = workbook["Sheet1" ]
      number_of_rows = worksheet.max_row
      number_of_columns = worksheet.max_column
      replacementTextKeyPairs ={'{}'.format(org_wrd):'{}'.format(rlc_wrd)}
      for i in range(number_of_columns): 
         for k in range(number_of_rows): 
            cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value) 
            for key in replacementTextKeyPairs.keys(): 
               if str(cellValue) == key: 
                  newCellValue = replacementTextKeyPairs.get(key) 
                  worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue) 
      workbook.save('o{}'.format(filename))
      new_path1 = os.path.abspath("o" + filename)
      
      data = open(new_path1, 'rb').read()
      #base64_encoded = base64.b64encode(data).decode('UTF-8')
      #output = base64_encoded.encode('utf-8')
      #worksheet = bytes(worksheet, 'utf-8')
      
   return render_template("xlsx-replace.html",op=f"o{filename}")
  #return render_template("xlsx-replace.html",op='output.xlsx')
 
if __name__ == '__main__':
   app.run(debug=True,port=80)